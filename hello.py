import os
from openpyxl import load_workbook

# 样本路径
sample_path = r'样本目录'

# 模版路径，注，只能使用xlsx后缀的模版。
templete_file = r'模板目录\水样全分析模板.xlsx'

# 输出文件夹下文件名
def getFile(dir):
    res = []
    for root, directory, files in os.walk(dir):
        for filename in files:
            name, suf = os.path.splitext(filename)
            res.append(filename)
    return res

files = getFile(sample_path)
print('文件名列表：\n',files)

# 将样本值存储在字典里面，即各个表中G4的值。如果G4和G5的值不同，则需要重新写判断条件。
dict = {}
for file in files:
    sample = os.path.join(sample_path,file)
    wb = load_workbook(sample)
    sheet = wb['sheet']
    name, suf = os.path.splitext(file)
    dict.update({name: sheet["G4"].value})

print('样本值：\n',dict)
# 将字典里的值赋到模板中
wb = load_workbook(templete_file)

# 模版的sheet名字
sheet = wb['95']

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if sheet.cell(row = i, column = j).value == 'As':
            sheet.cell(row = i, column = j+2).value = dict.get('As')
        elif sheet.cell(row = i, column = j).value == '总硬度':
            sheet.cell(row = i, column = j+2).value = dict.get('总硬度')
        elif sheet.cell(row = i, column = j).value == '总碱度':
            sheet.cell(row = i, column = j+2).value = dict.get('总碱度')
        elif sheet.cell(row = i, column = j).value == '暂硬度':
            sheet.cell(row = i, column = j+2).value = dict.get('暂硬度')
        elif sheet.cell(row = i, column = j).value == '永硬度':
            sheet.cell(row = i, column = j+2).value = dict.get('永硬度')
        elif sheet.cell(row = i, column = j).value == '负硬度':
            sheet.cell(row = i, column = j+2).value = dict.get('负硬度')
        elif sheet.cell(row = i, column = j).value == '游离CO2':
            sheet.cell(row = i, column = j+2).value = dict.get('游离CO2')
        elif sheet.cell(row = i, column = j).value == '溶解性总固体':
            sheet.cell(row = i, column = j+2).value = dict.get('溶解性总固体')
        elif sheet.cell(row = i, column = j).value == '耗氧量':
            sheet.cell(row = i, column = j+2).value = dict.get('耗氧量')
        elif sheet.cell(row = i, column = j).value == 'Cd':
            sheet.cell(row = i, column = j+2).value = dict.get('Cd')
        elif sheet.cell(row = i, column = j).value == 'Cl-':
            sheet.cell(row = i, column = j+2).value = dict.get('Cl')
        elif sheet.cell(row = i, column = j).value == 'CO32-':
            sheet.cell(row = i, column = j+2).value = dict.get('CO3')
        elif sheet.cell(row = i, column = j).value == 'F-':
            sheet.cell(row = i, column = j+2).value = dict.get('F')
        elif sheet.cell(row = i, column = j).value == 'Fe2+':
            sheet.cell(row = i, column = j+2).value = dict.get('Fe2')
        elif sheet.cell(row = i, column = j).value == 'Fe3+':
            sheet.cell(row = i, column = j+2).value = dict.get('Fe3')
        elif sheet.cell(row = i, column = j).value == 'H2PO4-':
            sheet.cell(row = i, column = j+2).value = dict.get('H2PO4')
        elif sheet.cell(row = i, column = j).value == 'H2SiO3':
            sheet.cell(row = i, column = j+2).value = dict.get('H2SiO3')
        elif sheet.cell(row = i, column = j).value == 'Ca+':
            sheet.cell(row = i, column = j+2).value = dict.get('Ca')
        elif sheet.cell(row = i, column = j).value == 'Mn':
            sheet.cell(row = i, column = j+2).value = dict.get('Mn')
        elif sheet.cell(row = i, column = j).value == 'Sr':
            sheet.cell(row = i, column = j+2).value = dict.get('Sr')
        elif sheet.cell(row = i, column = j).value == 'HCO3-':
            sheet.cell(row = i, column = j+2).value = dict.get('HCO3')
        elif sheet.cell(row = i, column = j).value == 'K+':
            sheet.cell(row = i, column = j+2).value = dict.get('K')
        elif sheet.cell(row = i, column = j).value == 'Mg+':
            sheet.cell(row = i, column = j+2).value = dict.get('Mg')
        elif sheet.cell(row = i, column = j).value == 'NH4+':
            sheet.cell(row = i, column = j+2).value = dict.get('NH4')
        elif sheet.cell(row = i, column = j).value == 'Na+':
            sheet.cell(row = i, column = j+2).value = dict.get('Na')
        elif sheet.cell(row = i, column = j).value == 'NO2-':
            sheet.cell(row = i, column = j+2).value = dict.get('NO2')
        elif sheet.cell(row = i, column = j).value == 'NO3-':
            sheet.cell(row = i, column = j+2).value = dict.get('NO3')
        elif sheet.cell(row = i, column = j).value == 'Hg':
            sheet.cell(row = i, column = j+2).value = dict.get('Hg')
        elif sheet.cell(row = i, column = j).value == 'OH-':
            sheet.cell(row = i, column = j+2).value = dict.get('OH')
        elif sheet.cell(row = i, column = j).value == 'Pb':
            sheet.cell(row = i, column = j+2).value = dict.get('Pb')
        elif sheet.cell(row = i, column = j).value == 'PH':
            sheet.cell(row = i, column = j+2).value = dict.get('Ph')
        elif sheet.cell(row = i, column = j).value == 'SO4-':
            sheet.cell(row = i, column = j+2).value = dict.get('SO4')
        elif sheet.cell(row = i, column = j).value == 'Zn':
            sheet.cell(row = i, column = j+2).value = dict.get('Zn')
        elif sheet.cell(row = i, column = j).value == 'Cr':
            sheet.cell(row = i, column = j+2).value = dict.get('Cr')
        elif sheet.cell(row = i, column = j).value == '侵蚀性CO2':
            sheet.cell(row = i, column = j+2).value = dict.get('侵蚀性CO2')
        elif sheet.cell(row = i, column = j).value == 'Cu':
            sheet.cell(row = i, column = j+2).value = dict.get('Cu')
        elif sheet.cell(row = i, column = j).value == '侵蚀性CO2':
            sheet.cell(row = i, column = j+2).value = dict.get('侵蚀性CO2')

# 保存
wb.save('样本分析结果.xlsx')
